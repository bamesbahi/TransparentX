from asyncio import sleep
from platform import release
from turtle import position
from pynput import keyboard
from pynput.keyboard import Key, Controller
#getting access to procces list
import psutil
#random
import random  
import string 
#delay
import time
#ip_api
from requests import get
def api():
    return "298cf64e"
#ip_Converting
import socket, struct
#notification
import plyer.platforms.win.notification
from plyer import notification
#registry
from winreg import * 
#toclose_script   
import sys
#clipboard
import win32clipboard
#timelog
from datetime import datetime
#---------------------------------------------------------SCRIPT------------------------------------------------------------------
#----------------Notify-----------------
#Toast
notification.notify("Developped by B.M", "Warmup Started")
#--------------Autorisation-------------
public=api()
def checking():
    ip = get('https://api.ipify.org').text
    return ip
def hex2ip(hex_ip):
    addr_long = int(hex_ip,16)
    hex(addr_long)
    hex_ip = socket.inet_ntoa(struct.pack(">L", addr_long))
    return hex_ip
if checking()!=hex2ip(public):
    sys.exit()
else:
     keyboard = Controller()
     import openpyxl
     from pathlib import Path 
# Checking if Launcher process is running or not.
def checkProcessRunning(processName):
    for proc in psutil.process_iter():
        try:
            if processName.lower() in proc.name().lower():
                return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    return False;

if not checkProcessRunning('TransparentX Launcher'):
    print('Error: USE THE LAUNCHER')
    time.sleep(5)
    sys.exit()

#---------------------------------------

#----------------Orders-----------------
#Getting spam orders 
xlsx_file_spam = Path('.\orders\spam.xlsx')
wb_obj_spam = openpyxl.load_workbook(xlsx_file_spam)
spam_sheet = wb_obj_spam.active
spam_count=spam_sheet.max_row
print("Spam orders Loaded : ",spam_count)
#Getting inbox orders 
xlsx_file_inbox = Path('.\orders\inbox.xlsx')
wb_obj_inbox = openpyxl.load_workbook(xlsx_file_inbox)
inbox_sheet = wb_obj_inbox.active
inbox_count=inbox_sheet.max_row
print("inbox orders Loaded : ",inbox_count)
#Getting bulk orders
xlsx_file_bulk = Path('.\orders/bulk.xlsx')
wb_obj_bulk = openpyxl.load_workbook(xlsx_file_bulk)
bulk_sheet = wb_obj_bulk.active
bulk_count=bulk_sheet.max_row
inbox=bulk_sheet.cell(row=1, column=2).value
spam=bulk_sheet.cell(row=2, column=2).value
#---------------------------------------

#--------------Functions----------------
def state(st):
     p=False
     f = open("State.txt", "r")
     if f.read()!="PAUSED" :
         if st!=True:
             notification.notify("TransparentX", "STARTED",timeout=3)
             print("RESUMED")
             time.sleep(3)
             
             
             
     else:
         if p!=st:
             notification.notify("TransparentX", "PAUSED",timeout=5)
         print("PAUSED")
         time.sleep(8)
         p=True
         state(False)
def start_browser():
     #start the browser
     keyboard.press(Key.cmd_l)
     keyboard.press('r')
     keyboard.release('r')
     keyboard.release(Key.cmd_l)
     time.sleep(2)
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     time.sleep(5)
     state(True)
def privateWin():
    time.sleep(2)
    keyboard.press(Key.ctrl)
    keyboard.press(Key.shift)
    keyboard.press('n')
    keyboard.release('n')
    keyboard.release(Key.shift)
    keyboard.release(Key.ctrl)
    time.sleep(2)
    keyboard.type('https://accounts.google.com/signin')
    time.sleep(2)
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    time.sleep(5)
    state(True)    
def actions(email):
    if check_position(email)!=False:
         
         not_spam(email)
         
         state(True)
         
         time.sleep(1)
         if check_position(email)!=False:
             bulk_spam(spam)
            
         time.sleep(1)
         fix_loading()
             
         time.sleep(3)
         
         state(True)
         if check_position(email)!=False:
             reporting(email)
             
         time.sleep(1)
         fix_loading()
         
         state(True)
         bulk_inbox(email,inbox)
         
    logout()
def login_email(email,password):
    print("Login ---> ",email)
    #login
    time.sleep(2)
    keyboard.type(email)
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    time.sleep(3)
    keyboard.type(password)
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    time.sleep(10)
    state(True)
    #open a new tab
    keyboard.press(Key.ctrl)
    keyboard.press('t')
    keyboard.release(Key.ctrl)
    keyboard.release('t')
    time.sleep(2)
    state(True)
    #closing the old tab
    keyboard.press(Key.ctrl)
    keyboard.press(Key.tab)
    keyboard.release(Key.tab)
    keyboard.release(Key.ctrl)
    time.sleep(2)
    keyboard.press(Key.ctrl)
    keyboard.press("w")
    keyboard.release(Key.ctrl)
    keyboard.release("w")
    time.sleep(2)
    state(True)
    #open inbox
    keyboard.type('https://mail.google.com/mail/u/0/#inbox')
    time.sleep(2)
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    print(email," ---> Connected Successfuly!")
    time.sleep(15)
    state(True)
   
    actions(email)
def open_console():
    #open console
     keyboard.press(Key.ctrl)
     keyboard.press(Key.shift)
     keyboard.press('j')
     keyboard.release(Key.ctrl)
     keyboard.release(Key.shift)
     keyboard.release('j')
     time.sleep(3)
     keyboard.press(Key.tab)
     keyboard.release(Key.tab)
     time.sleep(2)
     state(True)
def logout():
     time.sleep(1)
     keyboard.press(Key.alt)
     keyboard.press(Key.f4)
     keyboard.release(Key.f4)
     keyboard.release(Key.alt)
     print('Logout')
     print('___________________')
     state(True)
def select_all():
    print("Select all")
    time.sleep(1)
    keyboard.press('*')
    keyboard.press('a')
    keyboard.release('a')
    keyboard.release('*')
    time.sleep(1)
    state(True)
def not_spam(email):
    print('moving messages to inbox...')
    #Getting orders one by one
    for spm in spam_sheet.rows:
         notspm_count=0
         #set messages number
         while notspm_count<spm[1].value and check_position(email)!=False:
             #search on from name
             state(True)
             print('search on from name')
             keyboard.type('/')
             time.sleep(1)
             keyboard.type(spm[0].value)
             time.sleep(1)
             keyboard.press(Key.enter)
             keyboard.release(Key.enter)
             state(True)
             time.sleep(5)
             state(True)
             if isExist_light()!=False:
                 #select first messages and open
                 print('select first messages and open')
                 keyboard.press('x')
                 keyboard.release('x')
                 keyboard.press(Key.enter)
                 keyboard.release(Key.enter)
                 time.sleep(2)
                 state(True)
                 #mark not spam
                 print('mark not spam')
                 keyboard.type('.')
                 time.sleep(1)
                 i=0
                 while i<4:
                     keyboard.press(Key.down)
                     keyboard.release(Key.down)
                     time.sleep(1)
                     i=i+1
                 time.sleep(1)
                 keyboard.press(Key.enter)
                 keyboard.release(Key.enter)    
                 time.sleep(3)
                 state(True)
                 notspm_count=notspm_count+1 
             else:
                 print('No message found, skipping...! ---> ',spm[0].value)
                 break
    print('Done!')     
def reporting(email):
    #Getting orders one by one
    print('Reporting')
    for ibx in inbox_sheet.rows:
        #search on from name
         print('search on from name')
         keyboard.type('/')
         time.sleep(1)
         keyboard.type(ibx[0].value)
         time.sleep(1)
         keyboard.press(Key.enter)
         keyboard.release(Key.enter)
         time.sleep(5)
         state(True)
         #select first messages and open
         if isExist_light()!=False:
             keyboard.press('x')
             keyboard.release('x')
             keyboard.press(Key.enter)
             keyboard.release(Key.enter)
             time.sleep(2)
             messages_nbr=0
             state(True)
             while messages_nbr<ibx[1].value:
                 #star message
                 print('Star')
                 keyboard.type('s')
                 time.sleep(1)
                 #archive message
                 print('Archive')
                 keyboard.type('e')     
                 #reply
                 if ibx[2].value=="True" and isExist_light()!=False:
                     print('reply')
                     time.sleep(1)
                     keyboard.type('r')
                     time.sleep(2)
                     keyboard.type(specific_string(random.randint(5, 18)))
                     keyboard.type(" ")
                     keyboard.type(specific_string(random.randint(5, 18)))
                     time.sleep(1)
                     keyboard.press(Key.tab)
                     keyboard.release(Key.tab)
                     time.sleep(1)
                     keyboard.press(Key.space)
                     keyboard.release(Key.space)
                     time.sleep(5)
                     keyboard.press(Key.esc)
                     keyboard.release(Key.esc)
                     time.sleep(1)
                 keyboard.type('j')
                 messages_nbr=messages_nbr+1
                 time.sleep(2)
                 state(True)
    print('Done!')      
def specific_string(length):
    sample_string = 'pqrstuvwxy' # define the specific string  
    # define the condition for random string  
    result = ''.join((random.choice(sample_string)) for x in range(length))  
    return result
def isExist():
     #french----------------------
     # set clipboard data
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('aucun spam')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data_fr = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     time.sleep(1)
    #english----------------------
     # set clipboard data
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('no spam')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data_en = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     # clear clipboard value
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     if data_fr=="aucun spam" or data_en=="no spam":
         return False
     else:
         return True
def bulk_spam(spam):
     print('moving messages to inbox...')
     state(True)
     #Getting orders one by one
     while True:
         print('search on from name')
         keyboard.type('/')
         time.sleep(1)
         keyboard.type(spam)
         time.sleep(1)
         keyboard.press(Key.enter)
         keyboard.release(Key.enter)
         state(True)
         time.sleep(5)
         #state(True)
         if isExist()!=False:
             select_all()
             time.sleep(2)
             #state(True)
             #mark not spam
             keyboard.type(',')
             time.sleep(1)
             i=0
             while i<2:
                 keyboard.press(Key.tab)
                 keyboard.release(Key.tab)
                 time.sleep(1)
                 i=i+1
             time.sleep(1)
             keyboard.press(Key.enter)
             keyboard.release(Key.enter) 
             state(True)   
             time.sleep(3)
             #state(True)
         else:
             print('No message found, skipping...!')
             break
     print('Done!')  
def isExist_light():
     #french__1----------------------
     # set clipboard data
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('Aucun message')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data_fr = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     time.sleep(1)
    #english__1----------------------
     # set clipboard data
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('No messages')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data_en = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     #french__2----------------------
     # set clipboard data
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('corbeille ou marqués comme spam')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data_fr_2 = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     time.sleep(1)
    #english__2----------------------
     # set clipboard data
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('Trash or Spam')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data_en_2 = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     # clear clipboard value
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     if data_fr=="Aucun message" or data_en=="No messages" or data_fr_2=="corbeille ou marqués comme spam" or data_en_2=="Trash or Spam":
         return False
     else:
         return True  
def bulk_inbox(email,inbox):
     print('Final Step "SELECT ALL"...')
     #Getting orders one by one
     time.sleep(2)
     #click searchbox
     keyboard.type('/')
     time.sleep(2)

     #search on from name
     keyboard.type(inbox)
     time.sleep(3)
     keyboard.press(Key.enter)
     time.sleep(5)
     state(True)
     #state(True)
     while check_position(email)==True:
         #state(True)
         if isExist_light()!=False:
                 select_all()
                 time.sleep(1)
                 #set important
                 keyboard.type("+")
                 time.sleep(1)
                 state(True)
                 #read all 
                 keyboard.press(Key.shift)
                 keyboard.press('i')
                 keyboard.release('i')
                 keyboard.release(Key.shift)
                 state(True)
                 time.sleep(2)
                 #set star
                 keyboard.type('.')
                 time.sleep(2)
                 i=0
                 while i<3:
                     keyboard.press(Key.down)
                     time.sleep(1)
                     i=i+1
                 keyboard.press(Key.enter)
                 keyboard.release(Key.enter)
                 state(True)
                 print("deselect all")
                 time.sleep(1)
                 keyboard.press('*')
                 keyboard.press('n')
                 keyboard.release('n')
                 keyboard.release('*')
                 time.sleep(5)
                 state(True)
                 #click searchbox
                 keyboard.type('/')
                 time.sleep(2)
                 keyboard.press(Key.enter)
                 keyboard.release(Key.enter)
                 time.sleep(5)
                 state(True)
                 fix_loading()
         else:
             print('No message found, skipping...!')
             break
                  
     print('Done!')
def log_txt(data):
     now = datetime.now() # current date and time
     date_time = now.strftime("%m/%d/%Y, %H:%M:%S")
     file_object = open('C:/xampp/htdocs/log.txt', 'a')
     year = now.strftime("%Y")
     month = now.strftime("%m")
     day = now.strftime("%d")
     time = now.strftime("%H:%M:%S")
     fulldate=data+" ==> ["+year+"/"+month+"/"+day +" "+time+"]\n"
     file_object.write(fulldate)
     file_object.close()
def check_position(data):
     #french----------------------
     # set clipboard data
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('15 Go')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data_fr = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     time.sleep(1)
    #english----------------------
     # set clipboard data
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('15 GB')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data_en = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     # clear clipboard value
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     if data_fr=="15 Go" or data_en=="15 GB":
         return True
     else:
         print("check it please, skipping all tasks")
         error="check it please! =====> "+data
         log_txt(error) 
         return False 
def fix_loading():
     # clear clipboard value
     win32clipboard.OpenClipboard()
     win32clipboard.EmptyClipboard()
     win32clipboard.SetClipboardText('')
     win32clipboard.CloseClipboard()
     #checking
     keyboard.press(Key.ctrl)
     keyboard.type("f")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     keyboard.type('Loading...')
     keyboard.press(Key.enter)
     keyboard.release(Key.enter)
     keyboard.press(Key.esc)
     keyboard.release(Key.esc)
     time.sleep(1)
     keyboard.press(Key.ctrl)
     keyboard.type("c")
     keyboard.release(Key.ctrl)
     time.sleep(1)
     # get clipboard data
     win32clipboard.OpenClipboard()
     data = win32clipboard.GetClipboardData()
     win32clipboard.CloseClipboard()
     if data=="Loading...":
         keyboard.press(Key.f5)
         keyboard.release(Key.f5)
         time.sleep(10)
 
#Proxyswitcher
def proxy_changer(proxy,status):
     keyVal = 'Software\\Microsoft\\Windows\\CurrentVersion\\Internet Settings'
     key = OpenKey(HKEY_CURRENT_USER, keyVal, 0, KEY_ALL_ACCESS)
     SetValueEx(key, "ProxyServer", 0, REG_SZ, proxy)
     SetValueEx(key, "ProxyEnable", 0, REG_DWORD, status)
     CloseKey(key)
#---------------------------------------


#Getting Email List
xlsx_file = Path('.\seed.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

browser=False
proxy_changer("192.168.1.1",0)
#Process
for cl in sheet.rows:
     #Proxyswitcher
     print("Changing IP to : ",cl[2].value)
     proxy_changer(cl[2].value,1)
     time.sleep(15)
     if browser==False:
         print('start browser')
         start_browser()
         browser=True
         
     #private window
     print('start Private navigation')
     privateWin()
     time.sleep(8)
     #Start
     print('Process started')
     login_time="Login : "+cl[0].value
     log_txt(login_time)
     login_email(cl[0].value,cl[1].value)

     logout_time="Logout at: "
     log_txt(logout_time)
     proxy_changer(cl[2].value,0)
     print('--------End--------')


     

 

